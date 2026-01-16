from openpyxl import load_workbook
from openpyxl.styles import Font

workbook = load_workbook(filename='/home/bhillermann/Documents/Trade Analysis/Latest_trade.xlsx')

# Set the font format we want to use
default_font = Font(name='Rubik Light', size=10)

# Set the currenct formate
currency_format = '$#,##0.00'

# Iterate over all cells in all sheets and set the font
for x in workbook.sheetnames:
    sheet = workbook[x]
    for row in sheet["A:J"]:
        for cell in row:
            cell.font = default_font    

# Set the currency format on the summary table in SHU Data tab --------------
sheet = workbook['SHU Data']

# Definte which cells on the CMA pages need to be set to currency
for row in sheet["A:J"]:
    for cell in row:
        cell.number_format = currency_format


# Definte the CMA Summary pages we have to iterate through
cmas = ['Corangamite', 'Melbourne Water', 'Wimmera', 'Glenelg Hopkins', 
           'Goulburn Broken', 'West Gippsland', 'East Gippsland', 'Mallee', 
           'North Central', 'North East'
           ]

# Definte which cells on the CMA pages need to be set to currency
currency_cells = ('B3', 'B4', 'B5', 'B7', 'B8', 'B9', 'B10', 'B12')

# Iterate over the CMA sheets and set the currency format
for x in cmas:
    sheet = workbook[x]
    print(sheet.title)
    for cells in currency_cells:
        sheet[cells].number_format = currency_format
    
workbook.save(filename='/home/bhillermann/Documents/Trade Analysis/Latest_trade_updated.xlsx')