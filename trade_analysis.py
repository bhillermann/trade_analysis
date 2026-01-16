#!/usr/bin/env python3

from typing import Any
import pandas as pd
import numpy as np
from thefuzz import process
import copy
from datetime import datetime, timedelta
import argparse
from ghu_search import get_supply
from openpyxl import load_workbook
from openpyxl.styles import Font
import tempfile
import shutil
import time
import logging
import sys
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup, Tag
import xlsxwriter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

NVCR_URL = (
    "https://www.environment.vic.gov.au/"
    "native-vegetation/native-vegetation-removal-regulations"
)
URL_TEXT = "Traded credits information"


def wait_for_download(directory: str, timeout: int = 120) -> str:
    """Wait for .xlsx file to appear in directory after Selenium download."""
    start_time = time.time()
    logging.info(f"Waiting for download in: {directory}")

    last_part_file = None
    while time.time() - start_time < timeout:
        # Check for complete downloads
        xlsx_files = list(Path(directory).glob("*.xlsx"))
        if xlsx_files:
            file_path = str(xlsx_files[0])
            # Verify file is stable (size not changing)
            initial_size = Path(file_path).stat().st_size
            time.sleep(1)
            final_size = Path(file_path).stat().st_size
            if initial_size == final_size and final_size > 0:
                logging.info(f"Download complete: {file_path}")
                return file_path

        # Check for in-progress downloads
        part_files = list(Path(directory).glob("*.part"))
        if part_files:
            current_part = str(part_files[0])
            if current_part != last_part_file:
                logging.info(f"Download in progress: {current_part}")
                last_part_file = current_part

        time.sleep(0.5)

    # Timeout - log directory contents for debugging
    all_files = list(Path(directory).iterdir())
    logging.error(f"Download timeout. Directory contents: {[f.name for f in all_files]}")
    raise TimeoutError(f"Download did not complete within {timeout} seconds")


def _download_nvcr_file(tmpdir: str) -> str:
    """
    Internal helper to download NVCR trade data file using Selenium.
    Returns path to downloaded file in tmpdir.
    """
    options = Options()
    options.add_argument("--headless")
    options.set_preference("browser.download.folderList", 2)
    options.set_preference("browser.download.dir", tmpdir)
    options.set_preference("browser.download.useDownloadDir", True)
    options.set_preference("browser.helperApps.neverAsk.saveToDisk",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    driver = webdriver.Firefox(options=options)
    driver.set_page_load_timeout(60)

    try:
        # Load the NVCR page
        driver.get(NVCR_URL)
        time.sleep(5)  # Wait for page to fully render
        logging.info("Page loaded, searching for download link...")

        # Parse HTML to find download link
        html = driver.page_source
        soup = BeautifulSoup(html, "lxml")

        # Find and click the download link
        link_found = False
        for link in soup.find_all("a", href=True):
            if not isinstance(link, Tag):
                continue
            if URL_TEXT in link.get_text(strip=True):
                download_url = str(link.get("href"))
                logging.info(f"Download link found: {download_url}")

                # Click the link element (NOT driver.get - that blocks!)
                link_element = driver.find_element(By.CSS_SELECTOR, f'a[href="{download_url}"]')
                link_element.click()
                link_found = True
                break

        if not link_found:
            raise ValueError("Download link for traded credits not found.")

        # Wait for download to complete
        downloaded_file = wait_for_download(tmpdir)
        logging.info(f"File downloaded to: {downloaded_file}")
        return downloaded_file

    finally:
        driver.quit()


def get_trade_data() -> pd.ExcelFile:
    """Download NVCR trade data using Selenium with temporary file storage."""
    logging.info("Starting get_trade_data()")

    with tempfile.TemporaryDirectory() as tmpdir:
        downloaded_file = _download_nvcr_file(tmpdir)

        # Load into memory before temp directory cleanup
        excel_file = pd.ExcelFile(downloaded_file)
        return excel_file
    # tmpdir automatically deleted here


def save_nvcr_file(output_path: str) -> None:
    """Download NVCR trade data and save to specified path without analysis."""
    logging.info(f"Downloading NVCR trade data to: {output_path}")

    with tempfile.TemporaryDirectory() as tmpdir:
        downloaded_file = _download_nvcr_file(tmpdir)

        # Copy to user-specified location
        shutil.copy(downloaded_file, output_path)
        logging.info(f"NVCR trade data saved to: {output_path}")

# Call argparse and define the arguments
parser = argparse.ArgumentParser(description='Process trade prices and supply'
                                 'date to do trade analysis for the past '
                                 'year.')
parser.add_argument("-s", "--supply", required = False,
                    help='Name of the supply Excel data to read from. Default'
                    ' is to scrape new data.')
parser.add_argument("-i", "--input", required = False, 
                    help='The input trade price spreadsheet downloaded from '
                         'the NVCR. "https://www.environment.vic.gov.au/'
                         'native-vegetation/native-vegetation-removal-'
                         'regulations". '
                         'Not using this switch will download a new file.')
parser.add_argument("-o", "--output", default='Trade-Analysis.xlsx',
                    help='The name of the file you would like to write the '
                        'anlysis to. Default is "Trade-Analysis.xlsx" in the '
                        'current directory')
parser.add_argument("-b", "--start",
                    help='The date you wish to do the analysis from. '
                     'Default is 12 months ago')
parser.add_argument("-e", "--end",
                    help='The date you wish to do the analysis to. '
                        'Format is YYYY-MM-DD'
                        'Default is the end of the previous month')
parser.add_argument("--download-nvcr",
                    help='Download NVCR trade data file and save to specified '
                         'location without running analysis. Exits after download.')

args = parser.parse_args()

# Handle download-only mode first
if args.download_nvcr:
    print('Downloading NVCR trade data...')
    save_nvcr_file(args.download_nvcr)
    print(f'NVCR trade data saved as: {args.download_nvcr}')
    sys.exit(0)

# Import the Traded Credits data with pandas
# Define the Excel file to import
output_file = args.output

# Get supply data
try:
    if args.supply:
        print(f'Loading supply data from: {args.supply}')
        supply_df = pd.read_excel(args.supply, sheet_name=None)
    else:
        print('Downloading supply data...')
        supply_df = get_supply()
        print('Supply data downloaded.')
except Exception as e:
    print(f"Failed to get supply data: {e}")
    print("You can provide an existing supply file with --supply")
    sys.exit(1)

# Get trade data
try:
    if args.input:
        print(f'Loading trade data from: {args.input}')
        trade_df = pd.ExcelFile(args.input)
    else:
        print('Downloading NVCR trade data...')
        trade_df = get_trade_data()
        print('Trade data downloaded.')
except Exception as e:
    print(f"Failed to get trade data: {e}")
    print("You can provide an existing trade file with --input")
    sys.exit(1)


# Define the property IDs of the Water Authorities
wa = {}
wa['Corangamite'] = ['BBA-2252']
wa['Glenelg Hopkins'] = ['TFN-C0228 ']
wa['Melbourne Water'] = [
     'BBA-0277', 'BBA-0670', 'BBA-0677', 'BBA-0678']
wa['West Gippsland'] = ['BBA-3049', 'BBA-2845', 'BBA-2839', 'BBA-2790',
                        'BBA-2789', 'BBA-2751', 'BBA-2766', 'BBA-2623']

# trade_df and supply_df are already loaded from above

# Grab the HU tab
hu_df = trade_df.parse('Trade Prices by HU')

# Keep only the first 12 columns (the rest are empty unnamed columns)
hu_df = hu_df.iloc[:, :12]

# Rename the columns to something usable
hu_df = hu_df.set_axis([
     'date', 'cma', 'sbv', 'ghu', 'lt', 'sbu', 'ghu_price', 'shu_price',
     'species', 'price_in_gst', 'price_ex_gst', 'unnamed'],
     axis=1
     )

# start_date = input('Start date: ')
# end_date = input ('End date: ')

current_date = datetime.today()
end_date = current_date.replace(day=1)
end_date = end_date - timedelta(days=1)
start_date = end_date - timedelta(days=365)

if args.start:
    start_date = datetime.strptime(args.start, '%Y-%m-%d')

if args.end:
    end_date = datetime.strptime(args.end, '%Y-%m-%d')

# Ensure all 'cma' entries are type string
hu_df['cma'] = hu_df['cma'].map(str)

# Change Date from datetime to date
hu_df['date']=hu_df['date'].dt.date

# Drop the last column because it's not needed
hu_df = hu_df.drop(['unnamed'], axis=1)

# Convert numeric columns to proper types (handle string data from Excel)
hu_df['ghu'] = pd.to_numeric(hu_df['ghu'], errors='coerce').fillna(0)
hu_df['ghu_price'] = pd.to_numeric(hu_df['ghu_price'], errors='coerce').fillna(0)
hu_df['sbu'] = pd.to_numeric(hu_df['sbu'], errors='coerce').fillna(0)
hu_df['shu_price'] = pd.to_numeric(hu_df['shu_price'], errors='coerce').fillna(0)
hu_df['price_ex_gst'] = pd.to_numeric(hu_df['price_ex_gst'], errors='coerce').fillna(0)

# Grab the SHUs from the HU dataframe
shu_df = hu_df[pd.notnull(hu_df['species'])]

# Drop all HU trades outside of the date range
hu_df = hu_df[((hu_df['date'] >= start_date.date()) & (hu_df['date'] <= end_date.date()))]

# Drop the SHU columns we don't need
shu_df = shu_df.drop(['cma', 'sbv', 'ghu', 'ghu_price'], axis=1)

# Set 1 year and 3 year date ranges
one_year = end_date.date() - timedelta(days=365)
three_year = end_date.date() - timedelta(days=1095)

# Drop all SHU trades outside of a one year period from end date
shu_df_1y = shu_df[((shu_df['date'] >= one_year) &
                 (shu_df['date'] <= end_date.date()))]

# Drop all SHU trades outside of a one year period from end date
shu_df_3y = shu_df[((shu_df['date'] >= three_year) &
                 (shu_df['date'] <= end_date.date()))]

# Function to generate SHU summary from a filtered DataFrame
def create_shu_summary(filtered_df: pd.DataFrame) -> dict[str, Any]:
    total_sbu = filtered_df['sbu'].sum()
    return {
        'Number of SHU trades': filtered_df.groupby(['date', 'shu_price']).sum(numeric_only=True)['sbu'].count(),
        'Total SHUs traded': total_sbu,
        'Total Value of SHU trades': filtered_df['price_ex_gst'].sum(),
        'Average Price per SHU': filtered_df['price_ex_gst'].sum() / total_sbu if total_sbu > 0 else np.nan,
        'SHU Floor Price': filtered_df['shu_price'].min(),
        'SHU Ceiling Price': filtered_df['shu_price'].max(),
        'SHU median price': np.median(filtered_df['shu_price'].unique()) if not filtered_df.empty else np.nan
    }

# Create summaries
shu_summary_1y = create_shu_summary(shu_df_1y)
shu_summary_3y = create_shu_summary(shu_df_3y)

# Optional: convert to DataFrames for nicer display or export
shu_summary_df_1y = pd.DataFrame(list(shu_summary_1y.items()), columns=['Description', 'Value'])
shu_summary_df_3y = pd.DataFrame(list(shu_summary_3y.items()), columns=['Description', 'Value'])

# Drop the SHU trades so we only have GHU trades
hu_df = hu_df[pd.isnull(hu_df['species'])]

# Drop the HU columns we don't need
hu_df = hu_df.drop(['sbu', 'shu_price', 'species'], axis=1)

# Replace all NaN values with 0
hu_df['lt'] = hu_df['lt'].fillna(0)
# Make sure all LTs are integers
hu_df['lt'] = hu_df['lt'].map(int)

cmas = ['Corangamite', 'Melbourne Water', 'Port Phillip and Westernport', 
           'Wimmera', 'Glenelg Hopkins', 'Goulburn Broken', 'West Gippsland', 
           'East Gippsland', 'Mallee', 'North Central', 'North East'
           ]

# Clean up all the inconsistancies in CMA names
def fix_cmas(row: pd.Series) -> str:
    result = process.extractOne(row['cma'], cmas)  # type: ignore[attr-defined]
    if result is None:
        return str(row['cma'])
    return str(result[0])

hu_df['cma'] = hu_df.apply(lambda row: fix_cmas(row), axis=1)
hu_df = hu_df.replace('Port Phillip and Westernport', 'Melbourne Water')

# This needs to be cleaned up. Use normal headers and then relable after 
# calculations
summary = {'description': [
                            'Total GHUs traded', 'Total market value',
                            'Average price per GHU',
                            'Median price per GHU', 'Total GHUs without trees',
                            'Total value without trees',
                            'Average price without trees',
                            'Median price without trees', 'Floor price',
                            'Total LTs traded', 'Average LT value',
                            'Supply of Credits', 'Years of Supply', 
                            'LT Supply', 'Water Authority Supply (WA)',
                            'Years of Supply without WA'
                           ], 
                           'values': ['', '', '', '', '', '', '', '', 
                                      '', '', '', '', '', '', '', '']}

summary_df = pd.DataFrame(data=summary)

summaries: dict[str, pd.DataFrame] = {}

print('Calculating per CMA data-------------------------------------------\n')
for k, v in hu_df.groupby('cma'):
    cma_key = str(k)
    print(f'Crunching data for {cma_key}...\n')
    # Total GHUs traded
    summary_df.loc[0, 'values'] = v['ghu'].sum()
    # Total GHUs value
    summary_df.loc[1, 'values'] = v['price_ex_gst'].sum()
    # Average price per GHU
    summary_df.loc[2, 'values'] = v['price_ex_gst'].sum() / v['ghu'].sum()
    # Median price per GHU
    summary_df.loc[3, 'values'] = v['ghu_price'].median()
    # Total GHUs without trees
    summary_df.loc[4, 'values'] = v.loc[v['lt'] == 0].agg('ghu').sum()
    # Total value without trees
    summary_df.loc[5, 'values'] = v.loc[v['lt'] == 0].agg(
        'price_ex_gst').sum()
    # Average price without trees
    summary_df.loc[6, 'values'] = (
            v.loc[v['lt'] == 0].agg('price_ex_gst').sum()
            / v.loc[v['lt'] == 0].agg('ghu').sum()
        )
    # Median price without trees
    summary_df.loc[7, 'values'] = v.loc[v['lt'] == 0].agg(
         'ghu_price').median()
    # Floor price
    summary_df.loc[8, 'values'] = v['ghu_price'].min()
    # Total LTs traded
    summary_df.loc[9, 'values'] = v['lt'].sum()
    # Calculate the theoretical value of trees
    # (Total GHU value - ((Total GHUs - Total GHUs without trees)
    # * Avg price without trees) - Total value without trees)
    # / Total LTs Traded
    # Values are always numeric in this context, but pandas types them as Scalar
    val_1 = float(summary_df.at[1, 'values'])  # type: ignore[arg-type]
    val_0 = float(summary_df.at[0, 'values'])  # type: ignore[arg-type]
    val_4 = float(summary_df.at[4, 'values'])  # type: ignore[arg-type]
    val_5 = float(summary_df.at[5, 'values'])  # type: ignore[arg-type]
    val_6 = float(summary_df.at[6, 'values'])  # type: ignore[arg-type]
    val_9 = float(summary_df.at[9, 'values'])  # type: ignore[arg-type]
    summary_df.loc[10, 'values'] = (
        (val_1 - ((val_0 - val_4) * val_6) - val_5) / val_9
    )
    # Supply of Credits
    summary_df.loc[11, 'values'] = supply_df[cma_key].agg('GHU').sum()
    # Years of Supply
    val_11 = float(summary_df.loc[11, 'values'])  # type: ignore[arg-type]
    summary_df.loc[12, 'values'] = val_11 / val_0
    # LT Supply
    summary_df.loc[13, 'values'] = supply_df[cma_key].agg('LT').sum()
    # Calculate the number of credits owned by water authorities
    wa_credits = 0.0
    try:
        for x in wa[cma_key]:
            wa_credits = (wa_credits
                          + supply_df[cma_key].loc[supply_df[cma_key]['Credit Site ID']
                               == x].agg('GHU').sum())
    except KeyError:
        print(f"No Water Authority credits for {cma_key}.\n")
    # Water Authority Supply (WA)
    summary_df.loc[14, 'values'] = wa_credits
    # Years of Supply without WA
    summary_df.loc[15, 'values'] = (val_11 - wa_credits) / val_0



    summaries[cma_key] = copy.deepcopy(summary_df)


# Writing it all to Excel

print('Creating Excel Spreadsheet...\n\n')

writer = pd.ExcelWriter(output_file,
                    engine='xlsxwriter',
                    engine_kwargs={'options':{'strings_to_formulas': False}})

xlsx_workbook_raw = writer.book
assert xlsx_workbook_raw is not None, "ExcelWriter workbook should not be None with xlsxwriter engine"
xlsx_workbook: xlsxwriter.Workbook = xlsx_workbook_raw  # type: ignore[assignment]

# Define the different formats

currency_format = xlsx_workbook.add_format(
    {
        'num_format': '$#,##0.00'
    }
)

# Writing the HU information ------------------------------------------------
sheetname = 'HU Data'
# Write the HU dataframe to sheet HU Data
hu_df.to_excel(writer, sheet_name=sheetname, 
               startrow=1, header=False, index=False)

# Create some human readable headers
header = ('Date', 'CMA', 'SBV', 'GHU', 'LT', 'GHU Price', 
             'Price (in GST)', 'Price (ex GST)') 

column_settings = [{"header": column} for column in header]

# Get the dimensions of the dataframe.
(max_row, max_col) = hu_df.shape

# Set the active sheet to HU Data
worksheet = writer.sheets[sheetname]

# Add the Excel table structure. Pandas added the data.
worksheet.add_table(0, 0, max_row, max_col - 1, 
                    {
                        'columns': column_settings,
                        'style': 'Table Style Light 11',
                        'banded_columns': True
                    })

worksheet.set_column(max_col-3, max_col - 1, None, currency_format)

worksheet.autofit()

# End HU dataframe ----------------------------------------------------------

# Start - Writing SHU information to file -----------------------------------
sheetname = 'SHU Data'
# Write the SHU dataframe to sheet SHU Data
shu_df.to_excel(writer, sheet_name=sheetname, startrow=1, 
                header=False, index=False)

# Create some human readable headers
header = ('Date', 'LT',	'SHUs',	'SHU Price', 'Species', 
              'Price (in GST)', 'Price (ex GST)') 

column_settings = [{"header": column} for column in header]

# Get the dimensions of the dataframe.
(max_row, max_col) = shu_df.shape

# Set the active sheet to SHU Data
worksheet = writer.sheets[sheetname]

# Add the Excel table structure. Pandas added the data.
worksheet.add_table(0, 0, max_row, max_col - 1, 
                    {
                        'columns': column_settings,
                        'style': 'Table Style Light 11',
                        'banded_columns': True
                    })

# Set currency format on pricing columns
worksheet.set_column(max_col-2, max_col - 1, None, currency_format)
worksheet.set_column(3, 3, None, currency_format)

# Get the dimensions of the 3 year SHU Summary dataframe.
(max_row, max_col) = shu_summary_df_3y.shape

# Write the 3 year SHU Summary data
shu_summary_df_3y.to_excel(writer, sheet_name=sheetname, 
                        startrow=1, startcol=8, index=False, header=False)

# Write the 1 year SHU Summary data
shu_summary_df_1y.to_excel(writer, sheet_name=sheetname, 
                        startrow=max_row + 3, startcol=8, index=False, header=False)

# Add the headings for 1 year and 3 year SHU summaries

worksheet.write(0, 8, '3 Year SHU Summary')
worksheet.write(max_row + 2, 8, '1 Year SHU Summary')

# Add the Excel table structure. Pandas added the data.
worksheet.add_table(1, 8, max_row, max_col + 8 - 1, 
                    {
                        'style': 'Table Style Light 18',
                        'autofilter': False,
                        'header_row': False,
                        'first_column': True
                    })

worksheet.add_table(max_row + 3, 8, 2 * max_row + 2, max_col + 8 - 1, 
                    {
                        'style': 'Table Style Light 18',
                        'autofilter': False,
                        'header_row': False,
                        'first_column': True
                    })

# Autofit columns
worksheet.autofit()
# End SHU dataframe ---------------------------------------------------------

# Overview Summary Dataframe ------------------------------------------------
sheetname = 'HU Summary'
# Create high level summary data
hu_summary = hu_df.groupby('cma', as_index=False).agg({
    'ghu': 'sum',
    'lt': 'sum',
    'price_ex_gst': 'sum',
    'ghu_price': [
        lambda x: x[hu_df.loc[x.index, 'lt'] == 0].min(), 
        lambda x: x[hu_df.loc[x.index, 'lt'] == 0].max(), 
        lambda x: x[hu_df.loc[x.index, 'lt'] == 0].mean(), 
        lambda x: x[hu_df.loc[x.index, 'lt'] == 0].median()
    ]
})

new_hu_summary = pd.DataFrame()
summary_columns = summaries['Corangamite']['description'].to_list()
summary_columns.insert(0, 'CMA')

for x in summaries:
    temp_series = summaries[x]['values']
    temp_series = pd.concat([pd.Series([x]), temp_series], ignore_index=True)
    new_hu_summary = pd.concat([new_hu_summary, pd.DataFrame(temp_series)
        .transpose()], ignore_index=True)

new_hu_summary.columns = summary_columns

new_hu_summary = (new_hu_summary[['Supply of Credits', 'LT Supply']])

print(new_hu_summary)

hu_summary["GHU Weighted Average"] = hu_summary['price_ex_gst']['sum'] / hu_summary['ghu']['sum']
hu_summary = pd.concat([hu_summary, new_hu_summary], axis=1)

print(hu_summary)

# Write it to Excel 
hu_summary.to_excel(writer, sheet_name=sheetname)

# Create some human readable headers
header = ('Index', 'CMA', 'GHUs', 'LTs', 'Total Value', 'GHU Floor Price',
                     'GHU Ceiling Price', 'GHU Mean', 'GHU Median', 
                     'GHU Weighted Average', 'Available GHUs', 'Avalable LTs') 

column_settings = [{"header": column} for column in header]

# Get the dimensions of the dataframe.
(max_row, max_col) = hu_summary.shape

# Set the active sheet to SHU Data
worksheet = writer.sheets[sheetname]

# Add the Excel table structure. Pandas added the data.
worksheet.add_table(0, 0, max_row, max_col, 
                    {
                        'columns': column_settings,
                        'style': 'Table Style Light 11',
                        'banded_columns': True
                    })

# Set currency format on pricing columns
worksheet.set_column(max_col - 4, max_col - 2, None, currency_format)

# Autofit columns
worksheet.autofit()

# End Overview Summary data -------------------------------------------------

for cma in summaries:
        summaries[cma].columns = ['Metric', 'Value']
        summaries[cma].to_excel(writer, sheet_name=cma, index=False)
        writer.sheets[cma].autofit()

        # Get the dimensions of the dataframe.
        (max_row, max_col) = summaries[cma].shape

        worksheet = writer.sheets[cma]

        # Create some human readable headers
        header = ('Metric', 'Value')
        column_settings = [{"header": column} for column in header]

        # Add the Excel table structure. Pandas added the data.
        worksheet.add_table(0, 0, max_row, max_col - 1, 
                            {
                                'columns': column_settings,
                                'style': 'Table Style Light 11',
                                'banded_columns': True,
                                'autofilter': False
                            })

writer.close()

# Complete the formatting that can't be done by XlsxWriter. Use openpyxl

print('Putting final touches on formatting...\n\n')

workbook = load_workbook(filename=output_file)

# Set the font format we want to use
default_font = Font(name='Rubik Light', size=10)

# Set the currenct formate
currency_format = '$#,##0.00'

# Iterate over all cells in all sheets and set the font
for x in workbook.sheetnames:
    sheet = workbook[x]
    for row in sheet["A:L"]:
        for cell in row:
            cell.font = default_font

# Set the currency format on the summary table in SHU Data tab --------------
sheet = workbook['SHU Data']

# Define which cells in the summary table need to be set to currency
for row in sheet["J4:J8"]:
    for cell in row:
        cell.number_format = currency_format

for row in sheet["J12:J16"]:
    for cell in row:
        cell.number_format = currency_format


# Update the cmas to remove Port Phillip and Westernport before we iterate  
# through summary pages
cmas.remove('Port Phillip and Westernport')

# Define which cells on the CMA pages need to be set to currency
currency_cells = ('B3', 'B4', 'B5', 'B7', 'B8', 'B9', 'B10', 'B12')

# Iterate over the CMA sheets and set the currency format
for x in cmas:
    sheet = workbook[x]
    for cells in currency_cells:
        sheet[cells].number_format = currency_format
    
workbook.save(filename=output_file)

print('Analyses complete.')
